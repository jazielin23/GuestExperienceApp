import streamlit as st
import pandas as pd
import matplotlib.pyplot as plt
import user_defined_functions as udf
import config
import io

from openpyxl import Workbook
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.utils.dataframe import dataframe_to_rows



def missing_value_treatment():
    st.write("")
    st.markdown("""
<div style='background-color: #f9f9f9; padding: 12px; border-radius: 8px; border-left: 4px solid #0D3512; box-shadow: 0 3px 8px rgba(0, 0, 0, 0.1);'>
    <p style='font-size: 14px; color: #333333; font-family: Arial, sans-serif;'>
        <strong style='color: #0D3512;'>Here, users should upload new survey data and the required features.</strong><br><br>
        <strong style='color: #0D3512;'>Input 1 :</strong> Rawdata<br>
        <strong style='color: #0D3512;'>Input 2 :</strong> Required Features<br><br>
        <strong style='color: #0D3512;'>Output :</strong> 
        <ul style="padding-left: 20px; font-size: 14px;">
            <li>Cleaned dataset with missing target values optionally removed</li>
            <li>Features with high missing values dropped based on selected thresholds</li>
            <li>Missing values in remaining features replaced with <code>1000</code> for consistency</li>
            <li>Downloadable Excel file with <code>Included_Features</code> and <code>Excluded_Features</code> tabs</li>
            <li>Option to modify and re-upload the feature list before proceeding</li>
        </ul>
    </p>
</div>
""", unsafe_allow_html=True)




    if 'required_features' not in st.session_state:
        st.session_state.required_features = None
    if 'new_codes_df' not in st.session_state:
        st.session_state.new_codes_df = None
    
    if 'new_df_col_names' not in st.session_state:
        st.session_state.new_df_col_names = None   
    
    if 'new_df_after_col_drop' not in st.session_state:
        st.session_state.new_df_after_col_drop = None
   
    col1, col2 = st.columns(2) 
    new_file = None  # Ensure it's defined before use
    with col1:
        new_file = st.file_uploader("Upload CSV Data File (.csv)", type=["csv"])
    with col2:
        required_features = st.file_uploader("Upload columns required for the  model (Excel)", type=["xlsx"])
        if required_features is not None:
            st.session_state.required_features = required_features
    st.divider()
    if new_file is not None:
        new_df = pd.read_csv(new_file)              
        new_df_col_names = new_df.columns ##added this DC
        st.session_state.new_df_col_names = new_df_col_names

        if  new_df is not None and st.session_state.required_features is not None:
            new_df_new_col = new_df.copy() ####new_df_new_col is the actual data that we copied from session
            new_df_new_col.columns = st.session_state.new_df_col_names
            required_features_df = pd.read_excel(st.session_state.required_features)
            required_features_df['Required_features'] = required_features_df['Feature'] ##added this, DC
            new_df_after_col_drop = new_df_new_col[required_features_df['Required_features'].tolist()] ## Required_features.xlsx, will concat the rows of each column, convert to a list and filtering only the reqd features in the actual data which is in new_df_new_col
            original_df = required_features_df.copy()
            top_df = original_df.copy()
            top_df['Feature'] = top_df['Feature'].astype(str) + "_TOP"
            top_df['Description'] = top_df['Description'].astype(str) + "_TOP"
            top2_df = original_df.copy()
            top2_df['Feature'] = top2_df['Feature'].astype(str) + "_TOP2"
            top2_df['Description'] = top2_df['Description'].astype(str) + "_TOP2"
            # Concatenate original + TOP + TOP2
            combined_df = pd.concat([original_df, top_df, top2_df], ignore_index=True)
            combined_df = combined_df.sort_values(by="Feature").reset_index(drop=True)
            # st.write(combined_df)
            st.session_state.required_features_df = combined_df
            st.session_state.new_df_after_col_drop = new_df_after_col_drop
    if st.session_state.new_df_after_col_drop is not None:
        st.write("<h1 style='color: #0D3512; font-size: 25px; text-align:left; font-weight: normal;'> New data is uploaded to the tool </h1>", unsafe_allow_html=True)    
        df_dupe = st.session_state.new_df_after_col_drop.copy()
        duplicates = df_dupe.columns[df_dupe.columns.duplicated()].tolist()
        if duplicates:
            print("Duplicate columns detected:", duplicates)
        df_dupe = df_dupe.loc[:, ~df_dupe.columns.duplicated()]
        st.session_state.new_df_after_col_drop = df_dupe.copy()  # update the cleaned df
        st.dataframe(df_dupe.head(15), use_container_width=True, height=250, hide_index=True)
        new_df_after_col_drop = st.session_state.new_df_after_col_drop 

         # --- MISSING VALUE PERCENTAGES ---
        missing_percent = df_dupe.isnull().mean() * 100
        missing_df = missing_percent.reset_index()
        missing_df.columns = ['Feature', 'Missing Value Percentage']
        missing_df['Missing Value Percentage'] = missing_df['Missing Value Percentage'].round(1)
        missing_df = missing_df.sort_values(by='Missing Value Percentage', ascending=False).reset_index(drop=True) # Sort descending by 'Percentage Missing'
        # st.write("<h1 style='color: #0D3512; font-size: 25px; text-align:left; font-weight: normal;'>Missing Values Percentage for Selected Features </h1>", unsafe_allow_html=True)    
        st.markdown("""
        <div style="
            background-color: #f0f8f5; 
            padding: 12px 16px;
            border-left: 6px solid #0D3512;
            border-radius: 6px;
            margin-bottom: 15px;
        ">
            <h3 style="
                color: #0D3512; 
                font-size: 22px; 
                font-weight: 500;
                margin: 0;
            ">
                Missing Values Percentage for Selected Features
            </h3>
        </div>
        """, unsafe_allow_html=True)
        st.dataframe(missing_df, use_container_width=True, hide_index=True, height=250)
        new_df_after_col_drop.fillna(1000, inplace=True)
        st.session_state.new_df_after_col_drop.fillna(1000, inplace=True)
        # st.write("""
        #     <span style="
        #         color: #d9534f;  /* Bootstrap red */
        #         font-weight: 700; 
        #         font-size: 18px; 
        #         background: #fcebea;
        #         padding: 4px 10px;
        #         border-radius: 20px;
        #         box-shadow: 0 0 12px #d9534faa;
        #         display: inline-block;
        #         ">
        #         ⚠️ Missing values in the selected features have been replaced with <strong>1000</strong> to ensure smooth processing and consistent encoding!!!!
        #     </span>
        #     """, unsafe_allow_html=True)



        # --- MISSING VALUE TREATMENT FOR TARGET FEATURE ---
        st.markdown("""
        <div style="
            background-color: #f0f8f5; 
            padding: 12px 16px;
            border-left: 6px solid #0D3512;
            border-radius: 6px;
            margin-bottom: 15px;
        ">
            <h3 style="
                color: #0D3512; 
                font-size: 22px; 
                font-weight: 500;
                margin: 0;
            ">
                Missing Value Treatment for Target
            </h3>
        </div>
        """, unsafe_allow_html=True)
        all_features = df_dupe.columns.tolist() # Get all available features (from the dataframe used after replacing missing values)
        selected_target = st.selectbox("Select the target variable:", options=all_features) # Dropdown to select the target variable
        target_missing_count = df_dupe[selected_target].isnull().sum()

        target_total = len(df_dupe)
        target_missing_pct_round = round((target_missing_count / target_total) * 100, 1)
        target_missing_pct = (target_missing_count / target_total) * 100
        st.markdown(f"""
            <table style="border-collapse: collapse; width: 50%; font-size:16px;">
            <tr style="background-color:#0D3512; color: white;">
                <th style="border: 1px solid #ddd; padding: 8px; text-align:left;">Metric</th>
                <th style="border: 1px solid #ddd; padding: 8px; text-align:left;">Value</th>
            </tr>
            <tr>
                <td style="border: 1px solid #ddd; padding: 8px;">Total Rows</td>
                <td style="border: 1px solid #ddd; padding: 8px; color:#1f77b4;">{target_total:,}</td>
            </tr>
            <tr>
                <td style="border: 1px solid #ddd; padding: 8px;">Rows with Missing Target Values ('{selected_target}')</td>
                <td style="border: 1px solid #ddd; padding: 8px; color:#d62728;">{target_missing_count} ({target_missing_pct_round}%)</td>
            </tr>
            </table>
            """, unsafe_allow_html=True)
        


        # Conditional prompt if missing values are present
        if target_missing_count > 0:
            st.markdown("""
                <div style="
                    background-color: #ffe6e6; 
                    border-left: 6px solid #ff4d4d; 
                    padding: 10px; 
                    color: #b30000; 
                    font-weight: bold;
                    border-radius: 4px;
                    margin-bottom: 10px;
                ">
                    Missing values detected in the target column.
                </div>
                """, unsafe_allow_html=True)

            
            drop_choice = st.radio(
                "Do you want to remove rows with missing target values?",
                options=["Yes", "No"],
                horizontal=True,
                index=None
            )
            if drop_choice == "Yes":
                df_dupe_copy = df_dupe.copy() #### df_dupe_copy has all missing value traget rows and missing value features,1000 is not there
                df_dupe = df_dupe[df_dupe[selected_target].notna()].reset_index(drop=True)   #### df_dupe has all missing value rows of target removed, other features missing is still there, 1000 is not there
                # st.session_state.new_df_after_col_drop = df_dupe.copy()  # Update in session state if needed
                # st.session_state.new_df_after_col_drop.fillna(1000, inplace=True) ### after removing rows, fill missing value with 1000
                new_df_after_col_drop = st.session_state.new_df_after_col_drop
                st.markdown(f"""
                    <div style="
                        color: #d62728; 
                        font-weight: bold; 
                        margin-bottom: 10px;
                    ">
                        Rows with missing values in the target were removed.
                    </div>
                """, unsafe_allow_html=True)
                st.markdown(f"""
                <table style="border-collapse: collapse; width: 60%; font-size:16px;">
                    <tr style="background-color:#0D3512; color: white;">
                        <th style="border: 1px solid #ddd; padding: 8px; text-align:left;">Metric</th>
                        <th style="border: 1px solid #ddd; padding: 8px; text-align:left;">Value</th>
                    </tr>
                    <tr>
                        <td style="border: 1px solid #ddd; padding: 8px;">Total Rows Before Removal</td>
                        <td style="border: 1px solid #ddd; padding: 8px; color:#1f77b4;">{target_total:,}</td>
                    </tr>
                    <tr>
                        <td style="border: 1px solid #ddd; padding: 8px;">Total Rows After Removing Missing Target Values</td>
                        <td style="border: 1px solid #ddd; padding: 8px; color:#1f77b4;">{len(df_dupe):,}</td>
                    </tr>
                </table>
                """, unsafe_allow_html=True)
                
                # --- Recalculate Missing Percentage BEFORE Removal ---
                missing_percent_before = df_dupe_copy.isnull().mean() * 100
                missing_df_before = missing_percent_before.reset_index()
                missing_df_before.columns = ['Feature', 'Missing % Before']
                missing_df_before['Missing % Before'] = missing_df_before['Missing % Before'].round(1)

                # --- Recalculate Missing Percentage AFTER Removal ---
                missing_percent_after = df_dupe.isnull().mean() * 100
                missing_df_after = missing_percent_after.reset_index()
                missing_df_after.columns = ['Feature', 'Missing % After']
                missing_df_after['Missing % After'] = missing_df_after['Missing % After'].round(1)

                # --- Merge Before & After ---
                missing_comparison_df = pd.merge(missing_df_before, missing_df_after, on='Feature', how='outer')
                # missing_comparison_df.fillna(0, inplace=True)
                missing_comparison_df = missing_comparison_df.sort_values(by='Missing % Before', ascending=False)
                st.markdown("""
                    <div style="
                        background-color: #f0f8f5; 
                        padding: 12px 16px;
                        border-left: 6px solid #0D3512;
                        border-radius: 6px;
                        margin-top: 20px;
                        margin-bottom: 10px;
                    ">
                        <h3 style="
                            color: #0D3512; 
                            font-size: 20px; 
                            font-weight: 500;
                            margin: 0;
                        ">
                            Missing Value Percentage After Target Row Removal
                        </h3>
                    </div>
                """, unsafe_allow_html=True)
                st.dataframe(missing_comparison_df, use_container_width=True, hide_index=True,height=250)

                ### filtering dropdown for missing percentage threshold
                threshold_options = [90, 80, 70, 60, 50, 40, 30, 20, 10]
                selected_thresholds = st.multiselect(
                    "Select missing value % thresholds to drop features (features with missing % >= any selected threshold will be dropped):",
                    options=threshold_options
                    # default=[8
                )
                if selected_thresholds:
                    # Find features meeting any of the selected thresholds
                    features_to_drop = missing_df_after.loc[
                        missing_df_after['Missing % After'] >= min(selected_thresholds), 'Feature'
                    ].tolist()
                    
                    # Drop features
                    df_dupe_filtered = df_dupe.drop(columns=features_to_drop, errors='ignore').reset_index(drop=True)

                    
                    st.session_state.new_df_after_col_drop = df_dupe_filtered.copy()
                    st.markdown(f"""
                        <div style="font-size: 16px; font-weight: 600; margin-bottom: 8px;">
                            Features to be dropped based on selected thresholds <span style="color: #0D3512;">{selected_thresholds}</span>:
                        </div>
                    """, unsafe_allow_html=True)

                    if features_to_drop:
                        features_str = ", ".join(features_to_drop)
                        st.markdown(f"""
                            <div style="
                                color: #d62728; 
                                font-weight: bold; 
                                font-size: 16px; 
                                background-color: #fee6e6; 
                                padding: 8px 12px; 
                                border-radius: 5px;
                                margin-bottom: 15px;
                                ">
                                {features_str}
                            </div>
                        """, unsafe_allow_html=True)
                        st.write("""
                            <span style="
                                color: #d9534f;  /* Bootstrap red */
                                font-weight: 700; 
                                font-size: 18px; 
                                background: #fcebea;
                                padding: 4px 10px;
                                border-radius: 20px;
                                box-shadow: 0 0 12px #d9534faa;
                                display: inline-block;
                                ">
                                ⚠️ Missing values in the other features have been replaced with <strong>1000</strong> to ensure smooth processing and consistent encoding!!!!
                            </span>
                            """, unsafe_allow_html=True)
                    else:
                        st.markdown("""
                            <div style="
                                font-style: italic; 
                                color: #555555; 
                                font-size: 14px; 
                                margin-bottom: 15px;
                            ">
                                No features meet the criteria.
                            </div>
                        """, unsafe_allow_html=True)


                    
                    included_features_df = pd.DataFrame({'Included Features': df_dupe_filtered.columns.tolist()})
                    excluded_features_df = pd.DataFrame({'Excluded Features': features_to_drop}) if features_to_drop else pd.DataFrame({'Excluded Features': ['None']})
                    output = io.BytesIO() # Create Excel in memory with two tabs
                    with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
                        included_features_df.to_excel(writer, index=False, sheet_name='Included_Features')
                        excluded_features_df.to_excel(writer, index=False, sheet_name='Excluded_Features')
                        # writer.save()
                    processed_data = output.getvalue()

                    # Download button
                    st.download_button(
                        label="Download Features List (Included & Excluded)",
                        data=processed_data,
                        file_name='features_list.xlsx',
                        mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                    )
                    st.markdown("""
                        <div style="margin-top: 15px; background-color: #f9f9f9; border-left: 5px solid #0D3512; padding: 12px; border-radius: 6px;">
                            <strong> Next Step:</strong><br>
                            If you're satisfied with the included/excluded features, click <span style="color: green;"><strong>Pass</strong></span> below.<br><br>
                            Otherwise, open the <strong>features_list.xlsx</strong> file, move any feature(s) from <code>Excluded_Features</code> to <code>Included_Features</code>, and re-upload the modified file to proceed with custom selection.
                        </div>
                    """, unsafe_allow_html=True)
                    edited_features_file = st.file_uploader("📤 Upload Modified Features List (features_list.xlsx)", type=["xlsx"])

                    if edited_features_file is not None:
                        try:
                            edited_included_df = pd.read_excel(edited_features_file, sheet_name='Included_Features')
                            included_columns = edited_included_df['Included Features'].dropna().tolist()

                            # Rebuild the filtered DataFrame
                            final_filtered_df = df_dupe[included_columns].copy()
                            final_filtered_df.fillna(1000, inplace=True)
                            st.session_state.new_df_after_col_drop = final_filtered_df

                            st.success("✅ Features updated based on your uploaded file. New filtered data is ready for next steps.")
                            st.dataframe(final_filtered_df.head(10), use_container_width=True, height=250)

                        except Exception as e:
                            st.error(f"❌ Error processing the uploaded file: {str(e)}")

                    # 7. Pass button
                    if st.button("✅ Pass (Continue with filtered features)"):
                        st.success("✔ Proceeding with current filtered feature set.")
                        st.session_state.new_df_after_col_drop.fillna(1000, inplace=True)
                        st.dataframe(st.session_state.new_df_after_col_drop.head(10), use_container_width=True, height=250)





                        

                    
                else:
                    features_to_drop = []
                    st.session_state.new_df_after_col_drop = df_dupe.copy() #### removed target missing rows
                    st.session_state.new_df_after_col_drop.fillna(1000, inplace=True) ## filled 1000 for missing value forother features

                



                ####### encoding- after removal of missing value target rows ###############
                # Ensure the data copy is safe to modify
                
            
            elif drop_choice == "No":
                st.info("Missing values in the target column were retained.")
                st.write("""
            <span style="
                color: #d9534f;  /* Bootstrap red */
                font-weight: 700; 
                font-size: 18px; 
                background: #fcebea;
                padding: 4px 10px;
                border-radius: 20px;
                box-shadow: 0 0 12px #d9534faa;
                display: inline-block;
                ">
                ⚠️ Missing values in the selected features have been replaced with <strong>1000</strong> to ensure smooth processing and consistent encoding!!!!
            </span>
            """, unsafe_allow_html=True)
                ##### encoding- if user donot remove missing value target rows ####

        else:
            # st.write(f"<span style='color:green'><strong>No missing values found in the selected target column.</strong></span>", unsafe_allow_html=True)
            
            st.markdown("""
                <div style="
                    background-color: #e6ffe6; 
                    border-left: 6px solid #2eb82e; 
                    padding: 10px; 
                    color: #267326; 
                    font-weight: bold;
                    border-radius: 4px;
                    margin-bottom: 10px;
                ">
                    No missing values found in the selected target column. Proceed with Data Integration
                </div>
                """, unsafe_allow_html=True)
            
            st.write("""
            <span style="
                color: #d9534f;  /* Bootstrap red */
                font-weight: 700; 
                font-size: 18px; 
                background: #fcebea;
                padding: 4px 10px;
                border-radius: 20px;
                box-shadow: 0 0 12px #d9534faa;
                display: inline-block;
                ">
                ⚠️ Missing values in the selected features have been replaced with <strong>1000</strong> to ensure smooth processing and consistent encoding!!!!
            </span>
            """, unsafe_allow_html=True)

            ##### encoding if there is no missing values ####
            
    else:
        st.write('<span style="color:red; font-size:25px; font-weight:normal;"> **** Please upload New data and Required Features files **** </span >', unsafe_allow_html=True)


def data_integration_coding():
    st.write()
    st.markdown("""
    <div style='
        background-color: #f4fdf9; 
        padding: 16px 20px; 
        border-radius: 10px; 
        border-left: 6px solid #0D3512; 
        box-shadow: 0 4px 12px rgba(0, 0, 0, 0.07);
        margin-bottom: 20px;
    '>
        <p style='
            font-size: 16px; 
            color: #222222; 
            font-family: "Segoe UI", Tahoma, Geneva, Verdana, sans-serif;
            margin: 0;
        '>
            <strong style='
                color: #0D3512;
                font-size: 17px;
            '>Encoded Output:</strong><br><br>
            <span style='line-height: 1.6;'>
                Here, we generate <strong>encoded values</strong> for the selected features.<br>
            </span>
        </p>
    </div>
    """, unsafe_allow_html=True)

    new_df_with_tops = st.session_state.new_df_after_col_drop.copy()
    current_cols = new_df_with_tops.columns.tolist()
    for feature in st.session_state.required_features_df['Required_features']:
        if feature in current_cols:
            top_col = f"{feature}_TOP"
            top2_col = f"{feature}_TOP2"
            new_df_with_tops[top_col] = new_df_with_tops[feature]
            new_df_with_tops[top2_col] = new_df_with_tops[feature]
    st.session_state.new_df_after_col_drop = new_df_with_tops
    ###### for new codes.xlsx #######
    new_df_for_codes = st.session_state.new_df_after_col_drop.copy() 
    reshaped = []
    for col in new_df_for_codes.columns:
        unique_vals = new_df_for_codes[col].dropna().unique()
        for val in unique_vals:
            coded_value = 1 if val == 5 or val == "Excellent" or val == 5.0 else 0
            # If it's a _TOP column: only 5 or 5.0 → 1
            if col.endswith('_TOP'):
                coded_value = 1 if val in [5, 5.0] else 0
            # If it's a _TOP2 column: 4 or 5 → 1
            elif col.endswith('_TOP2'):
                coded_value = 1 if val in [4, 4.0, 5, 5.0] else 0
            reshaped.append({
                'Question Description': col,
                'Variable Rating Scale': val,
                'Coded Values': coded_value
            })
    new_codes_df = pd.DataFrame(reshaped)
    st.write("<h1 style='color: #FF0000; font-size: 20px; text-align:center; font-weight: normal;'> **** Below are the coded values for the features selected ****</h1>", unsafe_allow_html=True) 
    display_df = new_codes_df.copy()
    # display_df["Unique Values"] = display_df["Unique Values"].astype(str)
    # st.dataframe(display_df, use_container_width=True, hide_index=True)
    st.session_state.new_codes_df = new_codes_df  
    st.write("<h1 style='color: #0D3512; font-size: 25px; text-align:left; font-weight: normal;'> Modify the value in Coded Values column if required: </h1>", unsafe_allow_html=True)
    new_codes_df = st.data_editor(display_df.copy(), use_container_width=True, hide_index=True,disabled=("Question Description", "Variable Rating Scale"))
    st.session_state.new_codes_df = new_codes_df
    






def data_integration():
    st.write("<h1 style='color: #0D3512; font-size: 35px; text-align:left; font-weight: normal;'> Data Integration & Feature Selection </h1>", unsafe_allow_html=True)
 
#     st.markdown("""
# <div style='background-color: #f9f9f9; padding: 12px; border-radius: 8px; border-left: 4px solid #0D3512; box-shadow: 0 3px 8px rgba(0, 0, 0, 0.1);'>
#     <p style='font-size: 14px; color: #333333; font-family: Arial, sans-serif;'>
#         <strong style='color: #0D3512;'>Here, users should upload new survey data and the required features.</strong><br><br>
#         <strong style='color: #0D3512;'>Input 1 :</strong> Rawdata<br>
#         <strong style='color: #0D3512;'>Input 2 :</strong> Required Features<br><br>
#         <strong style='color: #0D3512;'>Output 1 :</strong> Filtered rawdata based on required features.<br>
#         <strong style='color: #0D3512;'>Output 2 :</strong> Encoded values for the features.
#     </p>
# </div>
# """, unsafe_allow_html=True)
    st.markdown("""
<div style='background-color: #f9f9f9; padding: 12px; border-radius: 8px; border-left: 4px solid #0D3512; box-shadow: 0 3px 8px rgba(0, 0, 0, 0.1);'>
    <p style='font-size: 14px; color: #333333; font-family: Arial, sans-serif;'>
        <strong style='color: #0D3512;'>Here, users should upload new survey data and the required features.</strong><br><br>
        <strong style='color: #0D3512;'>Input 1 :</strong> Rawdata<br>
        <strong style='color: #0D3512;'>Input 2 :</strong> Required Features<br><br>
        <strong style='color: #0D3512;'>Output :</strong> Encoded values for the features.
    </p>
</div>
""", unsafe_allow_html=True)



    if 'required_features' not in st.session_state:
        st.session_state.required_features = None
    if 'new_codes_df' not in st.session_state:
        st.session_state.new_codes_df = None
    
    if 'new_df_col_names' not in st.session_state:
        st.session_state.new_df_col_names = None   
    
    if 'new_df_after_col_drop' not in st.session_state:
        st.session_state.new_df_after_col_drop = None
    # create 2 columns for layout
    col1, col2 = st.columns(2) ## commented, DC
    new_file = None  # Ensure it's defined before use
    with col1:
        # new_file = st.file_uploader("Upload compressed CSV (.zip)", type=["zip"])  ##added
        # new_file = st.file_uploader("Upload SPSS Data File (.zsav)", type=["zsav"])
        new_file = st.file_uploader("Upload CSV Data File (.csv)", type=["csv"])

        

            



    with col2:
        required_features = st.file_uploader("Upload columns required for the  model (Excel)", type=["xlsx"])
        if required_features is not None:
            st.session_state.required_features = required_features


    st.divider()

    # if  new_file is not None:
    #     with zipfile.ZipFile(new_file, "r") as z:
    #         csv_filename = z.namelist()[0]
    #         with z.open(csv_filename) as f:
    #             new_df = pd.read_csv(f)

    if new_file is not None:
        new_df = pd.read_csv(new_file)

    # import pyreadstat
    # import tempfile
    # if new_file is not None:
    #     with tempfile.NamedTemporaryFile(delete=False, suffix=".zsav") as tmp:
    #         tmp.write(new_file.read())
    #         tmp_path = tmp.name

    #     # Read using pyreadstat
    #     new_df, meta = pyreadstat.read_sav(tmp_path)
                        



        new_df_col_names = new_df.columns ##added this DC
        st.session_state.new_df_col_names = new_df_col_names

        if  new_df is not None and st.session_state.required_features is not None:
            new_df_new_col = new_df.copy() ####new_df_new_col is the actual data that we copied from session
            new_df_new_col.columns = st.session_state.new_df_col_names
            required_features_df = pd.read_excel(st.session_state.required_features)
            required_features_df['Required_features'] = required_features_df['Feature'] ##added this, DC
            new_df_after_col_drop = new_df_new_col[required_features_df['Required_features'].tolist()] ## Required_features.xlsx, will concat the rows of each column, convert to a list and filtering only the reqd features in the actual data which is in new_df_new_col
            original_df = required_features_df.copy()
            top_df = original_df.copy()
            top_df['Feature'] = top_df['Feature'].astype(str) + "_TOP"
            top_df['Description'] = top_df['Description'].astype(str) + "_TOP"
            top2_df = original_df.copy()
            top2_df['Feature'] = top2_df['Feature'].astype(str) + "_TOP2"
            top2_df['Description'] = top2_df['Description'].astype(str) + "_TOP2"
            # Concatenate original + TOP + TOP2
            combined_df = pd.concat([original_df, top_df, top2_df], ignore_index=True)
            combined_df = combined_df.sort_values(by="Feature").reset_index(drop=True)
            # st.write(combined_df)
            st.session_state.required_features_df = combined_df
            st.session_state.new_df_after_col_drop = new_df_after_col_drop

            # print(new_df_after_col_drop)
    if st.session_state.new_df_after_col_drop is not None:
        st.write("<h1 style='color: #0D3512; font-size: 25px; text-align:left; font-weight: normal;'> New data is uploaded to the tool </h1>", unsafe_allow_html=True)    
        df_dupe = st.session_state.new_df_after_col_drop.copy()
        duplicates = df_dupe.columns[df_dupe.columns.duplicated()].tolist()
        if duplicates:
            print("Duplicate columns detected:", duplicates)
        df_dupe = df_dupe.loc[:, ~df_dupe.columns.duplicated()]
        st.session_state.new_df_after_col_drop = df_dupe.copy()  # update the cleaned df
        
        # Now show the dataframe without error
        st.dataframe(df_dupe.head(15), use_container_width=True, height=250, hide_index=True)

       
        
        # st.dataframe(st.session_state.new_df_after_col_drop.head(15),use_container_width=True,height=400,hide_index=True)  
        new_df_after_col_drop = st.session_state.new_df_after_col_drop 
        # csv_data = new_df_after_col_drop.to_csv(index=False)  # Convert dataframe to CSV
        # csv_buffer = io.StringIO(csv_data)  # Create an in-memory buffer

        # st.download_button(
        #     label="Download Selected Columns as CSV",
        #     data=csv_buffer.getvalue(),
        #     file_name="selected_columns.csv",
        #     mime="text/csv"
        # )

         # --- MISSING VALUE PERCENTAGES ---
        missing_percent = df_dupe.isnull().mean() * 100
        missing_df = missing_percent.reset_index()
        missing_df.columns = ['Feature', 'Missing Value Percentage']
        missing_df['Missing Value Percentage'] = missing_df['Missing Value Percentage'].round(1)
        missing_df = missing_df.sort_values(by='Missing Value Percentage', ascending=False).reset_index(drop=True) # Sort descending by 'Percentage Missing'
        # st.write("<h1 style='color: #0D3512; font-size: 25px; text-align:left; font-weight: normal;'>Missing Values Percentage for Selected Features </h1>", unsafe_allow_html=True)    
        st.markdown("""
        <div style="
            background-color: #f0f8f5; 
            padding: 12px 16px;
            border-left: 6px solid #0D3512;
            border-radius: 6px;
            margin-bottom: 15px;
        ">
            <h3 style="
                color: #0D3512; 
                font-size: 22px; 
                font-weight: 500;
                margin: 0;
            ">
                Missing Values Percentage for Selected Features
            </h3>
        </div>
        """, unsafe_allow_html=True)
        st.dataframe(missing_df, use_container_width=True, hide_index=True, height=250)
        new_df_after_col_drop.fillna(1000, inplace=True)
        st.write("""
            <span style="
                color: #d9534f;  /* Bootstrap red */
                font-weight: 700; 
                font-size: 18px; 
                background: #fcebea;
                padding: 4px 10px;
                border-radius: 20px;
                box-shadow: 0 0 12px #d9534faa;
                display: inline-block;
                ">
                ⚠️ Missing values in the selected features have been replaced with <strong>1000</strong> to ensure smooth processing and consistent encoding!!!!
            </span>
            """, unsafe_allow_html=True)



        # --- MISSING VALUE TREATMENT FOR TARGET FEATURE ---
        st.markdown("""
        <div style="
            background-color: #f0f8f5; 
            padding: 12px 16px;
            border-left: 6px solid #0D3512;
            border-radius: 6px;
            margin-bottom: 15px;
        ">
            <h3 style="
                color: #0D3512; 
                font-size: 22px; 
                font-weight: 500;
                margin: 0;
            ">
                Missing Value Treatment for Target
            </h3>
        </div>
        """, unsafe_allow_html=True)
        all_features = df_dupe.columns.tolist() # Get all available features (from the dataframe used after replacing missing values)
        selected_target = st.selectbox("Select the target variable:", options=all_features) # Dropdown to select the target variable
        target_missing_count = df_dupe[selected_target].isnull().sum()

        target_total = len(df_dupe)
        target_missing_pct_round = round((target_missing_count / target_total) * 100, 1)
        target_missing_pct = (target_missing_count / target_total) * 100
        st.markdown(f"""
            <table style="border-collapse: collapse; width: 50%; font-size:16px;">
            <tr style="background-color:#0D3512; color: white;">
                <th style="border: 1px solid #ddd; padding: 8px; text-align:left;">Metric</th>
                <th style="border: 1px solid #ddd; padding: 8px; text-align:left;">Value</th>
            </tr>
            <tr>
                <td style="border: 1px solid #ddd; padding: 8px;">Total Rows</td>
                <td style="border: 1px solid #ddd; padding: 8px; color:#1f77b4;">{target_total}</td>
            </tr>
            <tr>
                <td style="border: 1px solid #ddd; padding: 8px;">Rows with Missing Target Values ('{selected_target}')</td>
                <td style="border: 1px solid #ddd; padding: 8px; color:#d62728;">{target_missing_count} ({target_missing_pct_round}%)</td>
            </tr>
            </table>
            """, unsafe_allow_html=True)


        # Conditional prompt if missing values are present
        if target_missing_count > 0:
            st.markdown("""
                <div style="
                    background-color: #ffe6e6; 
                    border-left: 6px solid #ff4d4d; 
                    padding: 10px; 
                    color: #b30000; 
                    font-weight: bold;
                    border-radius: 4px;
                    margin-bottom: 10px;
                ">
                    Missing values detected in the target column.
                </div>
                """, unsafe_allow_html=True)

            
            drop_choice = st.radio(
                "Do you want to remove rows with missing target values?",
                options=["Yes", "No"],
                horizontal=True,
                index=None
            )
            if drop_choice == "Yes":
                df_dupe = df_dupe[df_dupe[selected_target].notna()].reset_index(drop=True)
                st.session_state.new_df_after_col_drop = df_dupe.copy()  # Update in session state if needed
                st.session_state.new_df_after_col_drop.fillna(1000, inplace=True) ### after removing rows, fill missing value with 1000
                new_df_after_col_drop = st.session_state.new_df_after_col_drop
                st.markdown(f"""
                    <div style="
                        color: #d62728; 
                        font-weight: bold; 
                        margin-bottom: 10px;
                    ">
                        Rows with missing values in the target were removed.
                    </div>
                """, unsafe_allow_html=True)
                st.markdown(f"""
                <table style="border-collapse: collapse; width: 60%; font-size:16px;">
                    <tr style="background-color:#0D3512; color: white;">
                        <th style="border: 1px solid #ddd; padding: 8px; text-align:left;">Metric</th>
                        <th style="border: 1px solid #ddd; padding: 8px; text-align:left;">Value</th>
                    </tr>
                    <tr>
                        <td style="border: 1px solid #ddd; padding: 8px;">Total Rows Before Removal</td>
                        <td style="border: 1px solid #ddd; padding: 8px; color:#1f77b4;">{target_total}</td>
                    </tr>
                    <tr>
                        <td style="border: 1px solid #ddd; padding: 8px;">Total Rows After Removing Missing Target Values</td>
                        <td style="border: 1px solid #ddd; padding: 8px; color:#1f77b4;">{len(df_dupe)}</td>
                    </tr>
                </table>
                """, unsafe_allow_html=True)


                ####### encoding- after removal of missing value target rows ###############
                # Ensure the data copy is safe to modify
                new_df_with_tops = new_df_after_col_drop.copy()
                for feature in st.session_state.required_features_df['Required_features']:
                    top_col = f"{feature}_TOP"
                    top2_col = f"{feature}_TOP2"
                    new_df_with_tops[top_col] = new_df_with_tops[feature]
                    new_df_with_tops[top2_col] = new_df_with_tops[feature]
                st.session_state.new_df_after_col_drop = new_df_with_tops
                ###### for new codes.xlsx #######
                new_df_for_codes = st.session_state.new_df_after_col_drop.copy() 
                reshaped = []
                for col in new_df_for_codes.columns:
                    unique_vals = new_df_for_codes[col].dropna().unique()
                    for val in unique_vals:
                        coded_value = 1 if val == 5 or val == "Excellent" or val == 5.0 else 0
                        # If it's a _TOP column: only 5 or 5.0 → 1
                        if col.endswith('_TOP'):
                            coded_value = 1 if val in [5, 5.0] else 0
                        # If it's a _TOP2 column: 4 or 5 → 1
                        elif col.endswith('_TOP2'):
                            coded_value = 1 if val in [4, 4.0, 5, 5.0] else 0
                        reshaped.append({
                            'Question Description': col,
                            'Unique Values': val,
                            'Coded Values': coded_value
                        })
                new_codes_df = pd.DataFrame(reshaped)
                st.write("<h1 style='color: #FF0000; font-size: 20px; text-align:center; font-weight: normal;'> **** Below are the coded values for the features selected ****</h1>", unsafe_allow_html=True) 
                display_df = new_codes_df.copy()
                # display_df["Unique Values"] = display_df["Unique Values"].astype(str)
                # st.dataframe(display_df, use_container_width=True, hide_index=True)
                st.session_state.new_codes_df = new_codes_df  
                st.write("<h1 style='color: #0D3512; font-size: 25px; text-align:left; font-weight: normal;'> Modify the value in Coded Values column if required: </h1>", unsafe_allow_html=True)
                new_codes_df = st.data_editor(display_df.copy(), use_container_width=True, hide_index=True,disabled=("Question Description", "Unique Values"))
                st.session_state.new_codes_df = new_codes_df
                
            
            elif drop_choice == "No":
                st.info("Missing values in the target column were retained.")
                

                
                

                ##### encoding- if user donot remove missing value target rows ####
                new_df_with_tops = new_df_after_col_drop.copy()
                for feature in st.session_state.required_features_df['Required_features']:
                    top_col = f"{feature}_TOP"
                    top2_col = f"{feature}_TOP2"
                    new_df_with_tops[top_col] = new_df_with_tops[feature]
                    new_df_with_tops[top2_col] = new_df_with_tops[feature]
                st.session_state.new_df_after_col_drop = new_df_with_tops
                ###### for new codes.xlsx #######
                new_df_for_codes = st.session_state.new_df_after_col_drop.copy() 
                reshaped = []
                for col in new_df_for_codes.columns:
                    unique_vals = new_df_for_codes[col].dropna().unique()
                    for val in unique_vals:
                        coded_value = 1 if val == 5 or val == "Excellent" or val == 5.0 else 0
                        # If it's a _TOP column: only 5 or 5.0 → 1
                        if col.endswith('_TOP'):
                            coded_value = 1 if val in [5, 5.0] else 0
                        # If it's a _TOP2 column: 4 or 5 → 1
                        elif col.endswith('_TOP2'):
                            coded_value = 1 if val in [4, 4.0, 5, 5.0] else 0
                        reshaped.append({
                            'Question Description': col,
                            'Unique Values': val,
                            'Coded Values': coded_value
                        })
                new_codes_df = pd.DataFrame(reshaped)
                st.write("<h1 style='color: #FF0000; font-size: 20px; text-align:center; font-weight: normal;'> **** Below are the coded values for the features selected ****</h1>", unsafe_allow_html=True) 
                display_df = new_codes_df.copy()
                # display_df["Unique Values"] = display_df["Unique Values"].astype(str)
                # st.dataframe(display_df, use_container_width=True, hide_index=True)
                st.session_state.new_codes_df = new_codes_df  
                st.write("<h1 style='color: #0D3512; font-size: 25px; text-align:left; font-weight: normal;'> Modify the value in Coded Values column if required: </h1>", unsafe_allow_html=True)
                new_codes_df = st.data_editor(display_df.copy(), use_container_width=True, hide_index=True,disabled=("Question Description", "Unique Values"))
                st.session_state.new_codes_df = new_codes_df



        else:
            # st.write(f"<span style='color:green'><strong>No missing values found in the selected target column.</strong></span>", unsafe_allow_html=True)
            st.markdown("""
                <div style="
                    background-color: #e6ffe6; 
                    border-left: 6px solid #2eb82e; 
                    padding: 10px; 
                    color: #267326; 
                    font-weight: bold;
                    border-radius: 4px;
                    margin-bottom: 10px;
                ">
                    No missing values found in the selected target column. 
                </div>
                """, unsafe_allow_html=True)

            ##### encoding if there is no missing values ####
            






        
            # Ensure the data copy is safe to modify
            new_df_with_tops = new_df_after_col_drop.copy()
            for feature in st.session_state.required_features_df['Required_features']:
                top_col = f"{feature}_TOP"
                top2_col = f"{feature}_TOP2"
                new_df_with_tops[top_col] = new_df_with_tops[feature]
                new_df_with_tops[top2_col] = new_df_with_tops[feature]
            st.session_state.new_df_after_col_drop = new_df_with_tops
            ###### for new codes.xlsx #######
            new_df_for_codes = st.session_state.new_df_after_col_drop.copy() 
            reshaped = []
            for col in new_df_for_codes.columns:
                unique_vals = new_df_for_codes[col].dropna().unique()
                for val in unique_vals:
                    coded_value = 1 if val == 5 or val == "Excellent" or val == 5.0 else 0
                    # If it's a _TOP column: only 5 or 5.0 → 1
                    if col.endswith('_TOP'):
                        coded_value = 1 if val in [5, 5.0] else 0
                    # If it's a _TOP2 column: 4 or 5 → 1
                    elif col.endswith('_TOP2'):
                        coded_value = 1 if val in [4, 4.0, 5, 5.0] else 0
                    reshaped.append({
                        'Question Description': col,
                        'Unique Values': val,
                        'Coded Values': coded_value
                    })
            new_codes_df = pd.DataFrame(reshaped)
            st.write("<h1 style='color: #FF0000; font-size: 20px; text-align:center; font-weight: normal;'> **** Below are the coded values for the features selected ****</h1>", unsafe_allow_html=True) 
            display_df = new_codes_df.copy()
            # display_df["Unique Values"] = display_df["Unique Values"].astype(str)
            # st.dataframe(display_df, use_container_width=True, hide_index=True)
            st.session_state.new_codes_df = new_codes_df  
            st.write("<h1 style='color: #0D3512; font-size: 25px; text-align:left; font-weight: normal;'> Modify the value in Coded Values column if required: </h1>", unsafe_allow_html=True)
            new_codes_df = st.data_editor(display_df.copy(), use_container_width=True, hide_index=True,disabled=("Question Description", "Unique Values"))
            st.session_state.new_codes_df = new_codes_df
    else:
        st.write('<span style="color:red; font-size:25px; font-weight:normal;"> **** Please upload New data and Required Features files **** </span >', unsafe_allow_html=True)



    

# def feature_categorization():
#     st.markdown("""
# <div style="background-color: #f9f9f9; padding: 12px; border-radius: 8px; box-shadow: 0 3px 8px rgba(0, 0, 0, 0.1);">
#     <h1 style="color: #0D3512; font-size: 28px; text-align: left; font-weight: normal;">Feature Categorization</h1>
#     <p style="font-size: 14px; color: #333333; font-family: Arial, sans-serif; line-height: 1.4;">
#         This page allows the user to bifurcate the features into explainers and segments as per the requirements.<br><br>
#         <strong style="color: #0D3512;">Explainers:</strong> Flag features as explainers if a feature explains the variability of the dependent variable.<br>
#         <strong style="color: #0D3512;">Segments:</strong> Flag features as segments if a feature can be used for clustering data or is not an explainer.
#     </p>
# </div>
# """, unsafe_allow_html=True)

#     ##classifiers are age,place, etc
#     ## explainers are ratings kind of things

#     if 'new_df_after_col_drop' in st.session_state:
#         data_fc = st.session_state.new_df_after_col_drop.copy()
#         print(data_fc.columns)
#     classifier_list = [col for col in data_fc.columns if any(desc in col for desc in config.col_descs)] ##changed
#     print(classifier_list)
#     explainer_list = set(data_fc.columns) - set(classifier_list)
    
#     df = pd.DataFrame({'Column Names': data_fc.columns})

#     def determine_flag(element):
#         if element in classifier_list:
#             return "segment"
#         elif element in explainer_list:
#             return "explainer"
#         else:
#             return None
    
#     df['Default Categorization'] = df['Column Names'].apply(determine_flag)
#     df['Custom Categorization'] = df['Column Names'].apply(determine_flag)
    
#     st.write("")
#     st.write("<h1 style='color: #0D3512; font-size: 25px; text-align:left; font-weight: normal;'> Modify the flag in Custom Categorization column if required: </h1>", unsafe_allow_html=True)
#     final_flags = st.data_editor(df.copy(), use_container_width=True, hide_index=True,disabled=("Column Names", "Default Categorization"))
#     st.session_state.final_flags = final_flags


#     if st.button("Update the above modifications"):
#         st.session_state.final_flags = final_flags
#         st.write("")
#         st.write("<h1 style='color: #0D3512; font-size: 25px; text-align:left; font-weight: normal;'> Updated Data: </h1>", unsafe_allow_html=True)
#         st.dataframe(st.session_state.final_flags, use_container_width=True, hide_index=True)
#         csv_data = st.session_state.final_flags.to_csv(index=False)
#         csv_buffer = io.StringIO(csv_data)

#         st.download_button(
#             label="Download Categorized Data as CSV",
#             data=csv_buffer.getvalue(),
#             file_name="updated_data.csv",
#             mime="text/csv"
#         )
#     else:
#         if 'final_flags' is None:
#             st.session_state.final_flags = df.copy()
#             st.dataframe(st.session_state.final_flags, use_container_width=True, hide_index=True)
#         else:
#             st.write("<h1 style='color: #FF0000; font-size: 20px; text-align:center; font-weight: normal;'> Press the button above. </h1>", unsafe_allow_html=True)
            

import streamlit as st
import pandas as pd
import io

def feature_categorization():
    st.markdown("""
    <div style="background-color: #f9f9f9; padding: 12px; border-radius: 8px; box-shadow: 0 3px 8px rgba(0, 0, 0, 0.1);">
        <h1 style="color: #0D3512; font-size: 28px; text-align: left; font-weight: normal;">Feature Categorization</h1>
        <p style="font-size: 14px; color: #333333; font-family: Arial, sans-serif; line-height: 1.4;">
            This page allows the user to bifurcate the features into explainers and segments as per the requirements.<br><br>
            <strong style="color: #0D3512;">Explainers:</strong> Flag features as explainers if a feature explains the variability of the dependent variable.<br>
            <strong style="color: #0D3512;">Segments:</strong> Flag features as segments if a feature can be used for clustering data or is not an explainer.
        </p>
    </div>
    """, unsafe_allow_html=True)

    if 'new_df_after_col_drop' in st.session_state:
        data_fc = st.session_state.new_df_after_col_drop.copy()
    else:
        st.error("No data found. Please upload or prepare the dataset first.")
        return

    # Define classifiers and explainers
    classifier_list = [col for col in data_fc.columns if any(desc in col for desc in config.col_descs)]
    explainer_list = set(data_fc.columns) - set(classifier_list)

    # Build the categorization DataFrame
    df = pd.DataFrame({'Column Names': data_fc.columns})

    def determine_flag(element):
        if element in classifier_list:
            return "segment"
        elif element in explainer_list:
            return "explainer"
        else:
            return None

    df['Default Categorization'] = df['Column Names'].apply(determine_flag)
    df['Custom Categorization'] = df['Column Names'].apply(determine_flag)

    # DROPDOWN TO FILTER BASED ON DEFAULT CATEGORIZATION
    st.write("### Select a feature category to view:")
    category_filter = st.selectbox("View features classified as:", options=["explainer", "segment"])

    filtered_df = df[df['Default Categorization'] == category_filter]

    if filtered_df.empty:
        st.warning(f"No features are classified as **{category_filter}** by default.")
    else:
        # st.markdown(f"These features are classified as **{category_filter}** by default. If you want to change their category, edit them in the 'Custom Categorization' column below.")
        st.markdown(f"""
<div style="background-color: #e8f5e9; padding: 12px 18px; border-radius: 6px; margin-top: 10px; border-left: 5px solid #2e7d32;">
    <p style="font-size: 15px; color: #1b5e20; font-family: Arial, sans-serif;">
        <strong>Note:</strong> These features are classified as <strong style="text-transform: capitalize;">{category_filter}</strong> by default.<br>
        If you want to change their category, edit them in the <em>'Custom Categorization'</em> column below.
    </p>
</div>
""", unsafe_allow_html=True)

        st.dataframe(filtered_df[['Column Names', 'Default Categorization']], use_container_width=True, hide_index=True)

    st.write("")
    st.markdown("<h1 style='color: #0D3512; font-size: 25px; text-align:left; font-weight: normal;'> Modify the flag in Custom Categorization column if required: </h1>", unsafe_allow_html=True)

    # Editable table for user to modify custom categorization
    final_flags = st.data_editor(df.copy(), use_container_width=True, hide_index=True,
                                 disabled=("Column Names", "Default Categorization"))
    st.session_state.final_flags = final_flags

    if st.button("Update the above modifications"):
        st.session_state.final_flags = final_flags
        st.write("")
        st.markdown("<h1 style='color: #0D3512; font-size: 25px; text-align:left; font-weight: normal;'> Updated Data: </h1>", unsafe_allow_html=True)
        st.dataframe(st.session_state.final_flags, use_container_width=True, hide_index=True)

        # CSV Download
        csv_data = st.session_state.final_flags.to_csv(index=False)
        csv_buffer = io.StringIO(csv_data)
        st.download_button(
            label="Download Categorized Data as CSV",
            data=csv_buffer.getvalue(),
            file_name="updated_data.csv",
            mime="text/csv"
        )
    else:
        if 'final_flags' not in st.session_state or st.session_state.final_flags is None:
            st.session_state.final_flags = df.copy()
        else:
            st.markdown("<h1 style='color: #FF0000; font-size: 20px; text-align:center; font-weight: normal;'> Press the button above. </h1>", unsafe_allow_html=True)








def data_encoding():
   
    st.markdown("""
<div style="background-color: #f9f9f9; padding: 12px; border-radius: 8px; box-shadow: 0 3px 8px rgba(0, 0, 0, 0.1);">
    <h1 style="color: #0D3512; font-size: 25px; text-align: left; font-weight: normal;">Data Encoding</h1>
    <p style="font-size: 14px; color: #333333; font-family: Arial, sans-serif; line-height: 1.5;">
        This page allows the user to define the codes corresponding to each question's choices as per the requirements.<br><br>
        <strong style="color: #0D3512;">- Columns flagged as segments</strong> in the 'Feature Categorization' page can be modified to use them for clustering.<br>
        <strong style="color: #0D3512;">- Columns flagged as explainers</strong> in the 'Feature Categorization' page, their choices can be recoded to suit the requirements.<br>
        <strong style="color: #0D3512;">- Sliders</strong> can be leveraged to filter low response rate questions.<br>
        <strong style="color: #0D3512;">- Pass button:</strong> If the codes mentioned in default meet the requirements, the user can hit the pass button.<br>
        <strong style="color: #0D3512;">- Modify codes:</strong> If the user wants to change the codes, they can download the file, change the values, and upload it again.
    </p>
</div>
""", unsafe_allow_html=True)


    st.divider()

    # took data from original file loaded and kept only selected columns and renamed columns as per final categorization
    ### from original data, for each column, created rows in Question description, unique values are the value like 24 is the age, count is the no.of people having that 24 age, distribution percentage is total count of that same qstn divided by 100, coded values is assigning a vlue to the unique values foex excellent i want to keep 1, good,vgood,poor i want to keep 0
    # st.write("st.session_state.final_flags",st.session_state.final_flags)
    classifier_flag = st.session_state.final_flags.loc[st.session_state.final_flags['Custom Categorization'] == "segment", 'Column Names'].to_list()
    explainer_flag = st.session_state.final_flags.loc[st.session_state.final_flags['Custom Categorization'] == "explainer", 'Column Names'].to_list()
    
    # st.write("st.session_state.new_df_after_col_drop",st.session_state.new_df_after_col_drop)
    classifier_df = st.session_state.new_df_after_col_drop.copy()
    classifier_df = classifier_df.loc[:,classifier_flag]
    transformed_classifier_df = udf.transform_data1(classifier_df) # transform_data function give the unique values for each column in df format
    transformed_classifier_df['Coded Values'] = transformed_classifier_df['Variable Rating Scale'] # created new column so that user can modify here
    st.write("<h1 style='color: #0D3512; font-size: 25px; text-align:left; font-weight: normal;'> Review the codes for segments (download and modify if required) </h1>", unsafe_allow_html=True)
    st.dataframe(transformed_classifier_df, use_container_width=True, hide_index=True)

    if st.button('Pass'):
        st.session_state.transformed_classifier_df = transformed_classifier_df
    else:
        st.write("<h1 style='color: #0D3512; font-size: 20px; text-align:left; font-weight: normal;'> If want to modify then download the above data , modify it and then upload in excel format, else press pass button above </h1>", unsafe_allow_html=True)
        upload_transformed_classifier = st.file_uploader("Upload modified segments file (Excel)", type=["xlsx"])

        if upload_transformed_classifier is not None:
            upload_transformed_classifier_df = pd.read_excel(upload_transformed_classifier)
            st.write("file uploaded")
            st.session_state.transformed_classifier_df = upload_transformed_classifier_df

# st.session_state.transformed_classifier_df data for
    st.divider()

    explainer_df = st.session_state.new_df_after_col_drop.copy()
    # st.write("explainer_df",explainer_df.dtypes)
    explainer_df = explainer_df.loc[:,explainer_flag]
    # print("explianer_df",explainer_df)
    transformed_explainer_df = udf.transform_data1(explainer_df)
    # st.write("***transformed_explainer_df",transformed_explainer_df)
    # st.write("############",st.session_state.new_codes_df)

    transformed_explainer_df['Question Description'] = transformed_explainer_df['Question Description'].astype(str)
    explainer_merged_df = transformed_explainer_df.merge(st.session_state.new_codes_df[['Question Description','Variable Rating Scale','Coded Values']], how='left', on=['Question Description','Variable Rating Scale'])
    explainer_merged_df['Coded Values'] = explainer_merged_df['Coded Values'].fillna(0)
    st.write("<h1 style='color: #0D3512; font-size: 25px; text-align:left; font-weight: normal;'> Review the distribution for explainers as per the original question choices </h1>", unsafe_allow_html=True)
    st.dataframe(explainer_merged_df, use_container_width=True, hide_index=True)

    codes = pd.concat([transformed_classifier_df,explainer_merged_df]).loc[:,['Question Description','Variable Rating Scale','Coded Values']]

    df_4_encode = st.session_state.new_df_after_col_drop.copy()
   
    transform_data1 = udf.transform_data(df_4_encode,codes).loc[:,explainer_flag]
    

    no_of_records1 = len(transform_data1)
    sum_summary1 =  transform_data1.sum()
    sum_summary1 = pd.to_numeric(sum_summary1, errors='coerce')
    # sum_summary1 = (sum_summary1/no_of_records1*100).reset_index().rename(columns={'index': 'Question Description', 0: 'Percentage of 1'}) ##changed 09-03-2025
    sum_summary1 = (sum_summary1/no_of_records1*100).reset_index().rename(columns={'index': 'Question Description', 0: 'Percentage of Top-Box'}).round(1)
    
    sum_summary1_copy=sum_summary1
    sum_summary1_copy=sum_summary1_copy.rename(columns={'Question Description':'Feature','Percentage of Top-Box':'Actual %'}).reset_index(drop=True)
    st.session_state.sum_summary1_copy = sum_summary1_copy
    # st.write("sum_summary1",sum_summary1_copy)
    

    ###### in coded-values we do the coding like excellent 1, others 0, then it may happen that in actual data some column became all 0's, so we decide, 30% values are 0 , dont take them
    st.write("<h1 style='color: #0D3512; font-size: 25px; text-align:left; font-weight: normal;'> Adjust the slider to choose the threshold </h1>", unsafe_allow_html=True)    

    percentage_of_1 = st.slider ('Review columns with low % of 1:' , min_value=0,max_value=100, value= 30, key='slider1')

    filtered_sum_summary1 = sum_summary1[sum_summary1['Percentage of Top-Box'] <= percentage_of_1]
    st.write("<h1 style='color: #0D3512; font-size: 25px; text-align:left; font-weight: normal;'> Filtered questions </h1>", unsafe_allow_html=True)    
    st.dataframe(filtered_sum_summary1, use_container_width=True, hide_index=True)

    # by considering above filtered columns 
    filtered_column1 = filtered_sum_summary1.iloc[:, 0].tolist()
    filter_explainer_merged_df = explainer_merged_df[explainer_merged_df['Question Description'].isin(filtered_column1)]
    not_filter_explainer_merged_df1 = explainer_merged_df[~explainer_merged_df['Question Description'].isin(filtered_column1)]
    st.write("<h1 style='color: #0D3512; font-size: 25px; text-align:left; font-weight: normal;'> Download and adjust the codes for filtered questions if required </h1>", unsafe_allow_html=True)    
    edited_new_codes_df = st.dataframe(filter_explainer_merged_df, use_container_width=True, hide_index=True)


    if st.button('Pass',key='pass1'):
        st.session_state.explainer_merged_df = explainer_merged_df
        st.session_state.final_file_4codes = pd.concat([st.session_state.transformed_classifier_df,st.session_state.explainer_merged_df])
    else:
        st.write("<h1 style='color: #0D3512; font-size: 20px; text-align:left; font-weight: normal;'> If want to modify then download the above data , modify it and then upload in excel format, else press pass button above </h1>", unsafe_allow_html=True)
        upload_edited_new_codes_file = st.file_uploader("Upload modified explainers file (Excel)", type=["xlsx"],key="uploader1")

        if upload_edited_new_codes_file is not None:
            st.write("file upload_edited_new_codes_df")
            st.session_state.upload_edited_new_codes_file = upload_edited_new_codes_file
            if 'upload_edited_new_codes_file' is not None:
                upload_edited_new_codes_df = pd.read_excel(st.session_state.upload_edited_new_codes_file)
                final_file_4codes = pd.concat([st.session_state.transformed_classifier_df,st.session_state.upload_edited_new_codes_df,not_filter_explainer_merged_df1])
                st.session_state.final_file_4codes = final_file_4codes
                st.write('final_file_4codes',final_file_4codes)    
        else:
            st.write("")

    if 'final_file_4codes' not in st.session_state:
        st.write("<h1 style='color: #FF0000; font-size: 20px; text-align:center; font-weight: normal;'> Press the pass button above or upload the modified files </h1>", unsafe_allow_html=True)
    else:
       


        transform_data_final = udf.transform_data(st.session_state.new_df_after_col_drop,st.session_state.final_file_4codes)

        st.session_state.transform_data_final = transform_data_final
        st.write("<h1 style='color: #0D3512; font-size: 25px; text-align:left; font-weight: normal;'> Final transformed data </h1>", unsafe_allow_html=True)
        st.dataframe(st.session_state.transform_data_final.head(20), use_container_width=True, hide_index=True)

    st.write("xxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxx      END      xxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxx")


# def data_summary():
#     st.write("<h1 style='color: #0D3512; font-size: 25px; text-align:left; font-weight: normal;'> New Data - Insights </h1>", unsafe_allow_html=True)
#     data = st.session_state.transform_data_final.copy()
#     for col in data.columns:
#         data[col] = pd.to_numeric(data[col], errors='coerce')

#     # data.dropna(inplace=True)
#     # Calculate correlation with 'Target' variable
#     st.write("<h1 style='color: #0D3512; font-size: 25px; text-align:left; font-weight: normal;'> Select the target column </h1>", unsafe_allow_html=True)
#     target = st.selectbox('', data.columns, label_visibility="visible")
#     correlation = data.corr()[target].drop(target).sort_values(ascending=False)

#     # Calculate value counts for all features except 'Target'
#     counts = data.drop(columns=[target]).apply(pd.Series.value_counts).fillna(0).astype(int)

#     # Calculate total number of rows in dataset
#     total_rows = len(data)

#     # Calculate response rates for each feature
#     response_rates = (counts.sum(axis=0) / total_rows) * 100

#     # Create summary DataFrame combining correlations and counts
#     summary_df = pd.DataFrame({
#         'Correlation': correlation,
#         'Response Rate (%)': response_rates
#     }).sort_values(by='Correlation', ascending=False)

#     st.write("<h1 style='color: #0D3512; font-size: 25px; text-align:left; font-weight: normal;'> Correlation with target (sorted based on correlation) </h1>", unsafe_allow_html=True)
#     st.dataframe(summary_df,use_container_width=True,height=500)
    
#     def plot_charts(feature):
#         if data[feature].count() > 0:
#             # Calculate value counts and percentages
#             value_counts = data[feature].value_counts()
#             percentages = (value_counts / value_counts.sum()) * 100
            
#             # Create DataFrame for table display
#             summary_df = pd.DataFrame({
#                 'Count': value_counts,
#                 'Percentage (%)': percentages.round(2)
#             })

#             subheader_html = f"""
#             <h2 style="color: blue; font-size: 20px; font-weight: normal;">{feature} Analysis</h2>
#             """

#             # Display the styled subheader
#             st.markdown(subheader_html, unsafe_allow_html=True)

#             fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(12, 6))
            
#             # Bar Chart
#             bars = data[feature].value_counts().plot(kind='bar', ax=ax1)
#             # ax1.bar_label(bars.containers[0])
#             ax1.set_xlabel('')
            

#             # Pie Chart
#             data[feature].value_counts().plot(kind='pie', autopct='%1.1f%%', ax=ax2)
            
#             st.pyplot(fig)


#     def display_table(feature):
#         if data[feature].count() > 0:
#             # Calculate value counts and percentages
#             value_counts = data[feature].value_counts()
#             percentages = (value_counts / value_counts.sum()) * 100
            
#             # Create DataFrame for table display
#             summary_df = pd.DataFrame({
#                 'Count': value_counts,
#                 'Percentage (%)': percentages.round(1)
#             })
#             subheader_html = f"""
#             <h2 style="color: blue; font-size: 20px; font-weight: normal;">{feature} Analysis</h2>
#             """
#             st.markdown(subheader_html, unsafe_allow_html=True)
#             st.dataframe(summary_df, use_container_width=True)
    
#     # Add dropdown to select a specific feature
#     st.write("<h1 style='color: #0D3512; font-size: 25px; text-align:left; font-weight: normal;'> Select a feature for detailed distribution </h1>", unsafe_allow_html=True)
#     selected_feature = st.selectbox('Select a feature:', correlation.index)
#     chart_type = st.radio("Select the below visualization option ", ('table', 'charts'))

#     # Plot or display table for the selected feature only
#     if chart_type == 'table':
#         display_table(selected_feature)
#     elif chart_type == 'charts':
#         plot_charts(selected_feature)

    

#     def export_all_features_to_excel(data, correlation):
#         output = io.BytesIO()
#         wb = Workbook()
#         ws = wb.active
#         ws.title = "Feature Summary"
#         row_num = 1  # Track current row in Excel
#         for feature in correlation.index:
#             if data[feature].count() == 0:
#                 continue
#             # Add feature title
#             ws.cell(row=row_num, column=1, value=f"{feature} Analysis")
#             row_num += 1
#             # Value counts and percentages
#             value_counts = data[feature].value_counts()
#             percentages = (value_counts / value_counts.sum()) * 100
#             feature_df = pd.DataFrame({
#                 'Count': value_counts,
#                 'Percentage (%)': percentages.round(2)
#             }).reset_index().rename(columns={'index': feature})
#             ## writing table to excel
#             for r in dataframe_to_rows(feature_df, index=False, header=True):
#                 for col_num, cell_val in enumerate(r, 1):
#                     ws.cell(row=row_num, column=col_num, value=cell_val)
#                 row_num += 1

#             ###create and save chart image
#             fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(10, 4))
#             value_counts.plot(kind='bar', ax=ax1, title='Bar Chart')
#             ax1.set_xlabel('')
#             value_counts.plot(kind='pie', autopct='%1.1f%%', ax=ax2)
#             ax2.set_ylabel('')
#             ax2.set_title('Pie Chart')

#             img_buf = io.BytesIO()
#             fig.savefig(img_buf, format='png', bbox_inches='tight')
#             plt.close(fig)
#             img_buf.seek(0)

#             img = OpenpyxlImage(img_buf) ## inserting image
#             img.anchor = f"A{row_num}"
#             ws.add_image(img)
#             row_num += 20  ### adding spacing after image

#         wb.save(output)
#         output.seek(0)
#         return output
     
#     excel_data = export_all_features_to_excel(data, correlation)
#     st.download_button(
#         label="Download All Feature Insights (Excel)",
#         data=excel_data,
#         file_name="feature_insights.xlsx",
#         mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
#     )

import streamlit as st
import pandas as pd
import matplotlib.pyplot as plt
import io
from openpyxl import Workbook
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.utils.dataframe import dataframe_to_rows

def data_summary():
    st.write("<h1 style='color: #0D3512; font-size: 25px; text-align:left; font-weight: normal;'> New Data - Insights </h1>", unsafe_allow_html=True)
    data = st.session_state.transform_data_final.copy()
    
    for col in data.columns:
        data[col] = pd.to_numeric(data[col], errors='coerce')

    # Select Target column
    st.write("<h1 style='color: #0D3512; font-size: 25px; text-align:left; font-weight: normal;'> Select the target column </h1>", unsafe_allow_html=True)
    target = st.selectbox('', data.columns, label_visibility="visible")

    # Calculate correlation with target
    correlation = data.corr()[target].drop(target).sort_values(ascending=False)

    # Calculate total number of rows in dataset
    total_rows = len(data)

    # Calculate value counts for all features except 'Target'
    counts = data.drop(columns=[target]).apply(pd.Series.value_counts).fillna(0).astype(int)
    response_rates = (counts.sum(axis=0) / total_rows) * 100

    summary_df = pd.DataFrame({
        'Correlation': correlation,
        'Response Rate (%)': response_rates
    }).sort_values(by='Correlation', ascending=False)

    st.write("<h1 style='color: #0D3512; font-size: 25px; text-align:left; font-weight: normal;'> Correlation with target (sorted based on correlation) </h1>", unsafe_allow_html=True)
    st.dataframe(summary_df, use_container_width=True, height=500)

    ## compute max count across all features
    max_count = 0
    for col in data.columns:
        if col == target:
            continue
        vc = data[col].value_counts()
        if not vc.empty:
            max_count = max(max_count, vc.max())

    def plot_charts(feature):
        if data[feature].count() > 0:
            value_counts = data[feature].value_counts()
            percentages = (value_counts / value_counts.sum()) * 100

            summary_df = pd.DataFrame({
                'Count': value_counts,
                'Percentage (%)': percentages.round(2)
            })

            subheader_html = f"""
            <h2 style="color: blue; font-size: 20px; font-weight: normal;">{feature} Analysis</h2>
            """
            st.markdown(subheader_html, unsafe_allow_html=True)

            fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(12, 6))

            bars = value_counts.plot(kind='bar', ax=ax1)
            ax1.set_ylim([0, max_count])  ##consistent bar height
            ax1.set_xlabel('')
            ax1.set_title('Bar Chart')

            value_counts.plot(kind='pie', autopct='%1.1f%%', ax=ax2)
            ax2.set_ylabel('')
            ax2.set_title('Pie Chart')

            st.pyplot(fig)

    def display_table(feature):
        if data[feature].count() > 0:
            value_counts = data[feature].value_counts()
            percentages = (value_counts / value_counts.sum()) * 100

            summary_df = pd.DataFrame({
                'Count': value_counts,
                'Percentage (%)': percentages.round(1)
            })

            subheader_html = f"""
            <h2 style="color: blue; font-size: 20px; font-weight: normal;">{feature} Analysis</h2>
            """
            st.markdown(subheader_html, unsafe_allow_html=True)
            st.dataframe(summary_df, use_container_width=True)

    # UI to select feature
    st.write("<h1 style='color: #0D3512; font-size: 25px; text-align:left; font-weight: normal;'> Select a feature for detailed distribution </h1>", unsafe_allow_html=True)
    selected_feature = st.selectbox('Select a feature:', correlation.index)
    chart_type = st.radio("Select the below visualization option ", ('table', 'charts'))

    if chart_type == 'table':
        display_table(selected_feature)
    elif chart_type == 'charts':
        plot_charts(selected_feature)

    def export_all_features_to_excel(data, correlation):
        output = io.BytesIO()
        wb = Workbook()
        ws = wb.active
        ws.title = "Feature Summary"
        row_num = 1

        for feature in correlation.index:
            if data[feature].count() == 0:
                continue

            ws.cell(row=row_num, column=1, value=f"{feature} Analysis")
            row_num += 1

            value_counts = data[feature].value_counts()
            percentages = (value_counts / value_counts.sum()) * 100

            feature_df = pd.DataFrame({
                'Count': value_counts,
                'Percentage (%)': percentages.round(2)
            }).reset_index().rename(columns={'index': feature})

            for r in dataframe_to_rows(feature_df, index=False, header=True):
                for col_num, cell_val in enumerate(r, 1):
                    ws.cell(row=row_num, column=col_num, value=cell_val)
                row_num += 1

            fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(10, 4))
            value_counts.plot(kind='bar', ax=ax1, title='Bar Chart')
            ax1.set_ylim([0, max_count])  ##consistent bar height
            ax1.set_xlabel('')
            ax2.set_title('Pie Chart')

            value_counts.plot(kind='pie', autopct='%1.1f%%', ax=ax2)
            ax2.set_ylabel('')

            img_buf = io.BytesIO()
            fig.savefig(img_buf, format='png', bbox_inches='tight')
            plt.close(fig)
            img_buf.seek(0)

            img = OpenpyxlImage(img_buf)
            img.anchor = f"A{row_num}"
            ws.add_image(img)

            row_num += 20

        wb.save(output)
        output.seek(0)
        return output

    # Download Button
    excel_data = export_all_features_to_excel(data, correlation)
    st.download_button(
        label="Download All Feature Insights (Excel)",
        data=excel_data,
        file_name="feature_insights.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


def new_feature_creation():
    st.markdown("""
<div style="background-color: #f9f9f9; padding: 12px; border-radius: 8px; box-shadow: 0 3px 8px rgba(0, 0, 0, 0.1);">
    <h1 style="color: #0D3512; font-size: 28px; text-align: left; font-weight: normal;">Feature Engineering</h1>
    <p style="font-size: 14px; color: #333333; font-family: Arial, sans-serif; line-height: 1.5;">
        Here, the user can create new variables. New variables can be created by:<br><br>
        <strong style="color: #0D3512;">- Adding</strong> two or more variables.<br>
        <strong style="color: #0D3512;">- Average</strong> between two or more variables.<br>
        <strong style="color: #0D3512;">- Maximum</strong> between two or more variables.<br>
        <strong style="color: #0D3512;">- Minimum</strong> between two or more variables.
    </p>
</div>
""", unsafe_allow_html=True)


   
    # st.session_state.data = st.session_state.transform_data_final.copy()
    if 'data' not in st.session_state:
        st.session_state.data = st.session_state.transform_data_final.copy()

    # Initialize session state
    if 'data' in st.session_state:
        # st.session_state.data = st.session_state.transform_data_final.copy()
        udf.display_custom_feature_engineering()
        st.write("<h1 style='color: #0D3512; font-size: 25px; text-align:left; font-weight: normal;'> Updated data based after adding new created feature </h1>", unsafe_allow_html=True)
        st.dataframe(st.session_state.data.head(15), use_container_width=True, hide_index=True)
        data = st.session_state.data
        

        # Convert DataFrame to CSV
        csv = data.to_csv(index=False)

        # Add download button
        st.download_button(
            label="📥 Download Data as CSV",
            data=csv,
            file_name='updated_data.csv',
            mime='text/csv'
        )

    else:
        st.write("<h1 style='color: #FF0000; font-size: 20px; text-align:center; font-weight: normal;'> **** Upload the necessary files in Data Integration page ****</h1>", unsafe_allow_html=True)


def run():
    st.divider()
    
    
    # if st.session_state.new_file is not None and st.session_state.new_codes is not None:
    sub_page = st.sidebar.radio(
        "Choose a sub task:",
        [ "Missing Value Treatment","Data Integration", "Feature Categorization", "Custom Encoding", "Data Summary", "New Feature Creation"]  # Added "Binning/Bucketing" to the list
    )
    if sub_page == "Missing Value Treatment":
        missing_value_treatment()
    if sub_page == "Data Integration":
        data_integration_coding()
    # elif sub_page == "Drop Columns":
    #     display_drop_columns()
    elif sub_page == 'Feature Categorization':
        feature_categorization()
    elif sub_page == "Custom Encoding":
        data_encoding()
    elif sub_page == "Data Summary":
        data_summary()
    elif sub_page == "New Feature Creation":
        new_feature_creation()

        st.markdown("### What's next?")
        col1, col2 = st.columns(2)

        with col1:
            if st.button("Go to Old Model Measurement"):
                st.session_state.page = "Old Model Measurement"
                st.rerun()

        with col2:
            if st.button("Skip to ML Engine"):
                st.session_state.page = "ML Engine"
                st.rerun()



####################


    # else:
    #     st.write('<span style="color:red; font-size:25px; font-weight:normal;"> **** Please upload Old data, New data and New Data Codes files on Data Integration page **** </span >', unsafe_allow_html=True)
